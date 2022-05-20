import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment

from product_to_xls import ProductToXls
from src.oz_goals import oz_goals
from src.wb_goals import wb_goals
from src.wb_terms import wb_terms, wb_categories
from src.oz_terms import oz_terms, oz_categories
from utilites import get_last_filename


class DataAnalytics:
    def __init__(self):
        self.json_filename = {'Ozon': get_last_filename('oz'), 'Wildberries': get_last_filename('wb')}
        self.goals = {'Ozon': oz_goals, 'Wildberries': wb_goals}
        self.current_goal = None
        self.platform_requests = {}
        self.first_rq_in_category = {}
        self.categories = {}
        self.date = {}
        self.rosel_goods = {}
        self.workbook = Workbook()
        self.shop = None
        self.sheet = None
        self.req_id = None

    def run(self):
        self.initiate_workbook()
        self.platform_actions(self.workbook.sheetnames[0])
        self.platform_actions(self.workbook.sheetnames[1])
        self.workbook.save(f'xls_result/result_{self.date["Wildberries"]}.xlsx')

    def platform_actions(self, platform):
        self.shop = platform
        self.read_requests()
        self.get_json()
        self.add_caption()
        self.add_titles()
        self.add_body()

    def get_json(self):
        filename = 'result/' + self.json_filename[self.shop]
        self.date[self.shop] = re.findall(r'\d{2}-\d{2}-202\d', self.json_filename[self.shop])[0]
        print(f'opening {filename}, by {self.date[self.shop]}')
        with open(filename, 'r', encoding='utf8') as file:
            self.rosel_goods = {}
            self.set_rosel_goods(file)

    def set_rosel_goods(self, file):
        sellers = ['РОСЭЛ', 'Фотон', 'Safeline', 'Контакт', 'КОНТАКТ Дом', 'Рекорд', 'ORGANIDE']
        for line in file:
            json_req = json.loads(line)
            req, req_id = list(json_req.values())[0], list(json_req)[0]
            products = []
            for order, prod in req.items():
                prod_id = prod['product']
                if prod['seller'] in sellers:
                    if self.check_incorrect_goods(req_id, prod_id):
                        continue
                    products.append(ProductToXls(prod_id, order, prod['name']))
            # products = [ProductToXls(p['product'], n, p['name']) for n, p in req.items() if p['seller'] in sellers]
            self.rosel_goods[req_id] = products
            # все бренды на 1-й странице
            # sls = [p['seller'] for n, p in req.items()]
            # print(self.platform_requests[req_id], sorted(set(sls)))

    def check_incorrect_goods(self, req_id, prod_id):
        ban_products_array = {'индикаторная отвертка': ['38710001', '79676933'],
                              'отвертка индикаторная': ['38710001', '79676933']
                              }
        incorrect_product_id_detected = False
        if self.platform_requests[req_id] in ban_products_array.keys():
            ban_list = ban_products_array[self.platform_requests.get(req_id, None)]
            if prod_id in ban_list:
                incorrect_product_id_detected = True
        return incorrect_product_id_detected

    def read_requests(self):
        terms, categories = [oz_terms, oz_categories] if self.shop == 'Ozon' else [wb_terms, wb_categories]
        self.platform_requests = {}
        self.first_rq_in_category = {}
        for cat_id, rq in terms.items():
            self.first_rq_in_category.update({list(rq)[0]: categories[cat_id]})
            self.platform_requests.update(rq)

    def add_caption(self):
        first_line = ['Отчет о представленности продукции по запросам на маркет-плейсах']
        second_line = ['Регион для которого осуществляется поисковая выдача - Москва.']
        self.sheet = self.workbook[self.shop]
        self.sheet.append(first_line)
        self.sheet.append([])
        self.sheet.append(second_line)
        self.sheet.append([])
        self.sheet['A1'].font = Font(name='Calibri', size=20, color="000000", bold=True)
        self.sheet.row_dimensions[1].height = 40
        self.sheet.column_dimensions['A'].width = 35
        self.sheet.column_dimensions['B'].width = 35

    def add_titles(self):
        date = self.date[self.shop].replace('-', '.')
        titles = ['Целевой запрос', 'Цель', date, 'Товары ' + date]
        self.sheet.append(titles)
        thick = Side(border_style="thick", color="000000")
        date_cell = self.sheet[f'C{self.sheet.max_row}']
        date_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        col_date = self.sheet.column_dimensions['C']
        col_goods = self.sheet.column_dimensions['D']
        col_date.width = 10
        col_goods.width = 80
        col_date.alignment = Alignment(horizontal='center')

    def add_body(self):
        goals_array = self.goals[self.shop]
        self.req_id = None
        for self.req_id in self.rosel_goods:
            self.check_category()
            self.current_goal = goals_array[self.req_id]
            self.set_body_lines()

    def set_body_lines(self):
        goal = self.current_goal
        goods = self.rosel_goods[self.req_id]
        rq = self.platform_requests[self.req_id]
        quantity_goods = len(goods)
        current_row_number = self.sheet.max_row + 1
        row_fin_number = None
        if quantity_goods:
            row_fin_number = current_row_number + quantity_goods - 1
            self.sheet.append([rq, goal, quantity_goods, goods[0].order_name()])
            for prod in goods[1:]:
                self.sheet.append([None, None, None, prod.order_name()])
        else:
            self.sheet.append([rq, goal, quantity_goods, None])
        self.merge_and_style_abc_columns(current_row_number, row_fin_number)
        self.check_goals(cell=f'C{current_row_number}', goods=goods)
        self.row_top_borders(current_row_number)

    def row_top_borders(self, row_number):
        thin = Side(border_style="thin", color="000000")
        row = self.sheet[row_number]
        for cell in row:
            cell.border = Border(top=thin)

    def merge_and_style_abc_columns(self, first_row, last_row):
        self.sheet[f'A{first_row}'].alignment = Alignment(horizontal='left', vertical='center')
        self.sheet[f'B{first_row}'].alignment = Alignment(horizontal='left', vertical='center')
        self.sheet[f'C{first_row}'].alignment = Alignment(horizontal='center', vertical='center')
        if last_row:
            self.sheet.merge_cells(f'A{first_row}:A{last_row}')
            self.sheet.merge_cells(f'B{first_row}:B{last_row}')
            self.sheet.merge_cells(f'C{first_row}:C{last_row}')

    def check_goals(self, cell, goods):
        cell = self.sheet[cell]
        cond = self.is_conditions_true(cell=cell, goods=goods)
        color_yes = 'C4D79B'
        color_no = 'E6B8B7'
        color = color_yes if cond else color_no
        cell.fill = PatternFill("solid", fgColor=color)

    def is_conditions_true(self, cell, goods):
        match self.current_goal:
            case 'на 1 странице, по популярности':
                return True if cell.value > 0 else False
            case 'не менее 5 SKU на 1 странице, по популярности':
                return True if cell.value >= 5 else False
            case 'не менее 4 SKU на 1 странице, по популярности':
                return True if cell.value >= 4 else False
            case 'не менее 2 SKU на 1 странице, по популярности':
                return True if cell.value >= 2 else False
            case 'не ниже 5 строки, по популярности':
                top5 = [prod.order for prod in goods if prod.order < 6]
                return True if top5 else False

    def check_category(self):
        if self.req_id not in self.first_rq_in_category.keys():
            return
        category_name = self.first_rq_in_category[self.req_id]
        self.sheet.append([category_name])
        max_row = self.sheet.max_row
        self.sheet.merge_cells(f'A{max_row}:D{max_row}')
        self.set_style_to_categoy_cell(max_row)

    def initiate_workbook(self):
        for shop in ['Ozon', 'Wildberries']:
            self.workbook.create_sheet(shop)
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        # ws['A1'].hyperlink = "http://www.google.com"

    def set_style_to_categoy_cell(self, row_order):
        cat_font = Font(name='Calibri', size=11, color="000000", bold=True)
        thin = Side(border_style="thin", color="000000")
        category_cell = self.sheet[f'A{row_order}']
        category_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        category_cell.fill = PatternFill("solid", fgColor="a6a6a6")
        category_cell.font = cat_font
        rd = self.sheet.row_dimensions[row_order]
        rd.height = 15


if __name__ == '__main__':
    res = DataAnalytics()
    res.run()
