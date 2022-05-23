import json
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment

from product_to_xls import ProductToXls
from src.oz_goals import oz_goals
from src.wb_goals import wb_goals
from src.wb_terms import wb_terms, wb_categories
from src.oz_terms import oz_terms, oz_categories
from utilites import last_month_json_files


class MonthlyDataAnalytics:
    def __init__(self):
        self.json_filenames = {'Ozon': last_month_json_files('oz'), 'Wildberries': last_month_json_files('wb')}
        self.goals = {'Ozon': oz_goals, 'Wildberries': wb_goals}
        self.all_terms = {'Ozon': oz_terms, 'Wildberries': wb_terms}
        self.all_categories = {'Ozon': oz_categories, 'Wildberries': wb_categories}
        self.current_goal = None
        self.platform_requests = {}
        self.platform_goods = []
        self.all_platform_goods = {}
        self.first_rq_in_category = {}
        self.categories = {}
        self.date = {}
        self.rosel_goods = {}
        self.workbook = Workbook()
        self.shop = None
        self.sheet = None
        self.req_id = None

    def run(self):
        self.initiate_workbook()
        self.platform_actions(self.workbook.sheetnames[0])
        self.platform_actions(self.workbook.sheetnames[1])
        last_date = list(self.all_platform_goods['Ozon'])[0]
        self.workbook.save(f"xls_result/{last_date.strftime('%B-%Y')}.xlsx")

    def platform_actions(self, platform):
        self.shop = platform
        self.read_requests()
        self.get_json()
        self.add_caption()
        self.add_titles()
        self.add_body()

    def get_json(self):
        json_files = self.json_filenames[self.shop]
        goods = {}
        for date, filename in json_files.items():
            print(f'opening {filename}')
            with open(filename, 'r', encoding='utf8') as file:
                self.rosel_goods = {}
                self.set_rosel_goods(file)
                goods.update({date: self.rosel_goods})
        self.rosel_goods = {}
        self.all_platform_goods[self.shop] = goods

    def set_rosel_goods(self, file):
        sellers = ['РОСЭЛ', 'Фотон', 'Safeline', 'Контакт', 'КОНТАКТ Дом', 'Рекорд', 'ORGANIDE']
        for line in file:
            json_req = json.loads(line)
            req, req_id = list(json_req.values())[0], list(json_req)[0]
            products = []
            for order, prod in req.items():
                prod_id = prod['product']
                if prod['seller'] in sellers:
                    if self.check_incorrect_goods(req_id, prod_id):
                        continue
                    products.append(ProductToXls(prod_id, order, prod['name']))
            self.rosel_goods[req_id] = products

    def check_incorrect_goods(self, req_id, prod_id):
        ban_prs = {'индикаторная отвертка': ['38710001', '79676933'], 'отвертка индикаторная': ['38710001', '79676933']}
        incorrect_product_id_detected = False
        if self.platform_requests[req_id] in ban_prs.keys():
            ban_list = ban_prs[self.platform_requests.get(req_id, None)]
            if prod_id in ban_list:
                incorrect_product_id_detected = True
        return incorrect_product_id_detected

    def read_requests(self):
        terms, categories = self.all_terms[self.shop], self.all_categories[self.shop]
        rqs, frc = {}, {}
        for cat_id, rq in terms.items():
            frc.update({list(rq)[0]: categories[cat_id]})
            rqs.update(rq)
        self.platform_requests, self.first_rq_in_category = rqs, frc

    def add_caption(self):
        first_line = ['Отчет о представленности продукции по запросам на маркет-плейсах']
        second_line = ['Регион для которого осуществляется поисковая выдача - Москва.']
        self.sheet = self.workbook[self.shop]
        self.sheet.append(first_line)
        self.sheet.append([])
        self.sheet.append(second_line)
        self.sheet.append([])
        self.sheet['A1'].font = Font(name='Calibri', size=20, color="000000", bold=True)
        self.sheet.row_dimensions[1].height = 40
        self.sheet.column_dimensions['A'].width = 35
        self.sheet.column_dimensions['B'].width = 35

    def add_titles(self):
        dates = [date.strftime('%d.%m.%Y') for date in self.all_platform_goods[self.shop]]
        products = ['Товары ' + date for date in dates]
        titles = ['Целевой запрос', 'Цель', *dates, *products]
        self.sheet.append(titles)
        self.add_titles_style(row_num=self.sheet.max_row, dates_quantity=len(dates))

    def add_titles_style(self, row_num, dates_quantity):
        sw = self.sheet
        thick = Side(border_style="thick", color="000000")
        first_date_column_number = ord('C')
        first_goods_column_number = first_date_column_number + dates_quantity
        for n in range(dates_quantity):
            date_column = chr(first_date_column_number + n)
            date_cell = sw[f'{date_column}{row_num}']
            date_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
            date_column_dim = sw.column_dimensions[date_column]
            date_column_dim.width = 10
            date_column_dim.alignment = Alignment(horizontal='center')
            goods_column = chr(first_goods_column_number + n)
            goods_column_dim = sw.column_dimensions[goods_column]
            goods_column_dim.width = 50

    def add_body(self):
        platform = self.shop
        self.platform_goods = [date for date in self.all_platform_goods[platform].values()]
        goals = self.goals[platform]
        self.req_id = None
        for self.req_id in self.platform_requests:
            self.check_category()
            self.current_goal = goals[self.req_id]
            self.set_body_lines()

    def set_body_lines(self):
        first_row_number = self.sheet.max_row + 1
        req_id = self.req_id
        all_dates_shop_products_for_rq = [prods[req_id] for prods in self.platform_goods]
        products_q = [len(el) for el in all_dates_shop_products_for_rq]
        max_goods_q = max(products_q)
        prods = [prs for prs in all_dates_shop_products_for_rq]
        rq = self.platform_requests[self.req_id]
        if max_goods_q:
            self.body_lines_to_table(products_q=products_q, prods=prods, max_goods_q=max_goods_q)
        else:
            empty = [None] * len(products_q)
            result = [rq, self.current_goal, *products_q, *empty]
            self.sheet.append(result)
        self.merge_body_cells(fst_row=first_row_number, lens=products_q)
        # self.row_top_borders(current_row_number)

    def body_lines_to_table(self, products_q, prods, max_goods_q):
        dates_quantity = len(products_q)
        for el in prods:
            len_rq_prod_list = len(el)
            if len_rq_prod_list < max_goods_q:
                for num in range(len_rq_prod_list, max_goods_q):
                    el.append(None)
        prod_lines = [[prods[date][n] for date in range(dates_quantity)] for n in range(max_goods_q)]
        for n, prl in enumerate(prod_lines):
            second = []
            for pq in products_q:
                if pq == 0:
                    second.append(0 if n > pq - 1 else pq)
                else:
                    second.append(None if n > pq - 1 else pq)
            prn = []
            for p in prl:
                prn.append(p.order_name()) if p else prn.append(None)
            result = [self.platform_requests[self.req_id], self.current_goal, *second, *prn]
            self.sheet.append(result)

    def merge_body_cells(self, fst_row, lens):
        max_q = max(lens)
        last_row = fst_row + max_q - 1 if max_q else fst_row
        # last_row = fst_row + max_q - 1
        self.merge_ab(fst_row, last_row)
        self.merge_goods_cells(first_row=fst_row, lens=lens)

    def merge_ab(self, first_row, last_row):
        sw = self.sheet
        sw[f'A{first_row}'].alignment = Alignment(horizontal='left', vertical='center')
        sw[f'B{first_row}'].alignment = Alignment(horizontal='left', vertical='center')
        if last_row:
            self.sheet.merge_cells(f'A{first_row}:A{last_row}')
            self.sheet.merge_cells(f'B{first_row}:B{last_row}')

    # quantity_goods = 0
    # current_row_number = self.sheet.max_row + 1
    # row_fin_number = None
    # if quantity_goods:
    #     row_fin_number = current_row_number + quantity_goods - 1
    #     self.sheet.append([rq, goal, quantity_goods, goods[0].order_name()])
    #     for prod in goods[1:]:
    #         self.sheet.append([None, None, None, prod.order_name()])
    # else:
    #     self.sheet.append([rq, goal, quantity_goods, None])
    # self.merge_and_style_abc_columns(current_row_number, row_fin_number)
    # self.check_goals(cell=f'C{current_row_number}', goods=goods)
    # self.row_top_borders(current_row_number)
    #

    def merge_goods_cells(self, first_row, lens):
        sw = self.sheet
        first_cell = ord('C')
        rows = max(lens)
        for n, l in enumerate(lens):  # 4 0 5
            column = chr(first_cell + n)
            cell = f'{column}{first_row}'
            if l:
                last_row = first_row + l - 1
                sw.merge_cells(f'{cell}:{column}{last_row}')
            else:
                last_row = first_row + rows - 1 if rows else first_row
                # last_row = first_row + rows - 1
                sw.merge_cells(f'{cell}:{column}{last_row}')
                column2 = chr(first_cell + n + len(lens))
                cell2 = f'{column2}{first_row}'
                sw.merge_cells(f'{cell2}:{column2}{last_row}')
            sw[cell].alignment = Alignment(horizontal='center', vertical='center')
            self.check_goals(cell, self.platform_goods[n][self.req_id])
            self.row_top_borders(first_row)

    def check_goals(self, cell, goods):
        cell = self.sheet[cell]
        cond = self.is_conditions_true(cell=cell, goods=goods)
        color_yes = 'C4D79B'
        color_no = 'E6B8B7'
        color = color_yes if cond else color_no
        cell.fill = PatternFill("solid", fgColor=color)

    def row_top_borders(self, row_number):
        thin = Side(border_style="thin", color="000000")
        row = self.sheet[row_number]
        for cell in row:
            cell.border = Border(top=thin)

    def is_conditions_true(self, cell, goods):
        match self.current_goal:
            case 'на 1 странице, по популярности':
                return True if cell.value > 0 else False
            case 'не менее 5 SKU на 1 странице, по популярности':
                return True if cell.value >= 5 else False
            case 'не менее 4 SKU на 1 странице, по популярности':
                return True if cell.value >= 4 else False
            case 'не менее 2 SKU на 1 странице, по популярности':
                return True if cell.value >= 2 else False
            case 'не ниже 5 строки, по популярности':
                top5 = []
                for prod in goods:
                    if prod:
                        if prod.order < 6:
                            top5.append(prod.order)
                # top5 = [prod.order for prod in goods if prod.order < 6]
                return True if top5 else False

    def check_category(self):
        if self.req_id not in self.first_rq_in_category.keys():
            return
        category_name = self.first_rq_in_category[self.req_id]
        self.sheet.append([category_name])
        max_row = self.sheet.max_row
        max_column = chr(ord('A') + self.sheet.max_column - 1)
        self.sheet.merge_cells(f'A{max_row}:{max_column}{max_row}')
        self.set_style_to_categoy_cell(max_row)

    def initiate_workbook(self):
        for shop in ['Ozon', 'Wildberries']:
            self.workbook.create_sheet(shop)
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def set_style_to_categoy_cell(self, row_order):
        cat_font = Font(name='Calibri', size=11, color="000000", bold=True)
        thin = Side(border_style="thin", color="000000")
        category_cell = self.sheet[f'A{row_order}']
        category_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        category_cell.fill = PatternFill("solid", fgColor="a6a6a6")
        category_cell.font = cat_font
        rd = self.sheet.row_dimensions[row_order]
        rd.height = 15


if __name__ == '__main__':
    res = MonthlyDataAnalytics()
    res.run()
