from pprint import pprint

from openpyxl import load_workbook, Workbook

from src.oz_terms import oz_categories
from src.wb_terms import wb_categories
from utilites import last_tables


class CombineTables:
    def __init__(self):
        self.platforms = ['Ozon', 'Wildberries']
        self.xls_table_filenames_dict = last_tables()
        self.tables = []
        self.workbook = Workbook()
        self.cur_sw = None
        self.sheet = None
        self.categories = {'Ozon': oz_categories.values(), 'Wildberries': wb_categories.values()}
        self.content = {'Ozon': [], 'Wildberries': []}

    def run(self):
        self.load_tables()
        self.initiate_workbook()
        self.table_actions()
        self.workbook.save(f"xls_result/{list(self.xls_table_filenames_dict.keys())[0].strftime('%B-%Y')}.xlsx")

    def load_tables(self):
        for name in self.xls_table_filenames_dict.values():
            print(f'loading {name}')
            self.tables.append(load_workbook(name))

    def initiate_workbook(self):
        for shop in self.platforms:
            self.workbook.create_sheet(shop)
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def table_actions(self):
        for platform in self.platforms:
            self.sheet = self.workbook[platform]
            self.set_caption(platform)
            for table in self.tables:
                self.cur_sw = table[platform]
                self.get_table_elements(platform)
            self.new_table_data(platform)

    def set_caption(self, shop):
        table = self.tables[-1][shop]
        for i, row in enumerate(table):
            num = i + 1
            if num == 2 or num == 4:
                self.sheet.append([])
            a, b = f'A{num}', f'B{num}'
            table_a, table_b = table[a].value, table[b].value
            if any([table_a, table_b]):
                self.sheet.append([table_a, table_b])

    def get_table_elements(self, platform):
        table = self.cur_sw
        categories = list(self.categories[platform])
        content = {}
        cat = None
        rq_and_goods, rq, goods = {}, None, []
        for i, row in enumerate(table):
            num = i + 1
            if num < 6:
                continue
            a, b, d = f'A{num}', f'B{num}', f'D{num}'
            cel_a, cel_b, cel_d = table[a].value, table[b].value, table[d].value
            if cel_a in categories:
                if rq_and_goods:
                    rq_and_goods[rq] = goods
                    content[cat] = rq_and_goods
                    rq_and_goods, rq, goods = {}, None, []
                cat = cel_a
                rq_and_goods = {}
            elif cel_a:
                if rq:
                    rq_and_goods[rq] = goods
                goods = []
                rq = cel_a
            if cel_d:
                goods.append(cel_d)
        content[cat] = rq_and_goods
        self.content[platform].append(content)
        print()

    def new_table_data(self, platform):
        categories = {}
        for category in self.categories[platform]:
            for lst in self.content[platform]:
                cat_cont = lst[category]
            categories = {lst[category] for lst in self.content[platform]}
            break


if __name__ == '__main__':
    combiner = CombineTables()
    combiner.run()
